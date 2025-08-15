import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Union, List, Dict, Any
import re

class ExcelConverter:
    """Converts extracted PDF data to Excel format with proper Chinese character support"""
    
    def __init__(self):
        self.default_font = Font(name='Microsoft YaHei', size=11)  # Good for Chinese characters
        self.header_font = Font(name='Microsoft YaHei', size=12, bold=True)
        self.header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        self.border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
    
    def convert_to_excel(self, data: Union[str, List[pd.DataFrame], Dict], 
                        extraction_method: str, original_filename: str) -> io.BytesIO:
        """
        Convert extracted data to Excel format
        
        Args:
            data: Extracted data from PDF
            extraction_method: Method used for extraction
            original_filename: Original PDF filename for reference
            
        Returns:
            BytesIO object containing Excel file
        """
        try:
            wb = Workbook()
            # Remove default sheet
            if wb.active:
                wb.remove(wb.active)
            
            if extraction_method == "Text Extraction":
                self._add_text_sheet(wb, data, original_filename)
            
            elif extraction_method == "Table Detection":
                self._add_table_sheets(wb, data, original_filename)
            
            else:  # Both Text and Tables
                if isinstance(data, dict):
                    if data.get("text"):
                        self._add_text_sheet(wb, data["text"], original_filename)
                    if data.get("tables"):
                        self._add_table_sheets(wb, data["tables"], original_filename)
                else:
                    # Fallback - treat as text
                    self._add_text_sheet(wb, data, original_filename)
            
            # If no sheets were added, create a summary sheet
            if not wb.worksheets:
                self._add_summary_sheet(wb, "No data extracted", original_filename)
            
            # Save to BytesIO
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file
            
        except Exception as e:
            # Create error sheet
            wb = Workbook()
            if wb.active:
                wb.remove(wb.active)
            self._add_error_sheet(wb, str(e), original_filename)
            
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file
    
    def _add_text_sheet(self, workbook: Workbook, text_data: Union[str, List[str]], filename: str):
        """Add text data to Excel sheet"""
        ws = workbook.create_sheet(title="Extracted Text")
        
        # Add header
        ws['A1'] = f"Text Extracted from: {filename}"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        
        # Process text data
        if isinstance(text_data, list):
            # Multiple pages
            row_num = 3
            for i, page_text in enumerate(text_data):
                ws[f'A{row_num}'] = f"Page {i+1}:"
                ws[f'A{row_num}'].font = Font(name='Microsoft YaHei', size=11, bold=True)
                row_num += 1
                
                # Split text into paragraphs
                paragraphs = page_text.split('\n')
                for paragraph in paragraphs:
                    if paragraph.strip():
                        ws[f'A{row_num}'] = paragraph.strip()
                        ws[f'A{row_num}'].font = self.default_font
                        ws[f'A{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
                        row_num += 1
                
                row_num += 1  # Extra space between pages
        else:
            # Single text block
            row_num = 3
            paragraphs = text_data.split('\n')
            for paragraph in paragraphs:
                if paragraph.strip():
                    ws[f'A{row_num}'] = paragraph.strip()
                    ws[f'A{row_num}'].font = self.default_font
                    ws[f'A{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
                    row_num += 1
        
        # Adjust column width
        ws.column_dimensions['A'].width = 80
        
        # Add timestamp
        ws[f'A2'] = f"Extracted on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(name='Microsoft YaHei', size=9, italic=True)
    
    def _add_table_sheets(self, workbook: Workbook, tables_data: List[pd.DataFrame], filename: str):
        """Add table data to Excel sheets"""
        if not tables_data:
            self._add_summary_sheet(workbook, "No tables found in PDF", filename)
            return
        
        for i, df in enumerate(tables_data):
            # Create sheet name
            sheet_name = f"Table_{i+1}"
            if hasattr(df, 'attrs') and 'page' in df.attrs:
                sheet_name = f"Table_{i+1}_Page_{df.attrs['page']}"
            
            # Ensure sheet name is valid (Excel limitation)
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', sheet_name)[:31]
            
            ws = workbook.create_sheet(title=sheet_name)
            
            # Add table header info
            ws['A1'] = f"Table {i+1} from: {filename}"
            ws['A1'].font = self.header_font
            ws['A1'].fill = self.header_fill
            
            if hasattr(df, 'attrs') and 'page' in df.attrs:
                ws['A2'] = f"Source: Page {df.attrs['page']}"
                ws['A2'].font = Font(name='Microsoft YaHei', size=9, italic=True)
                start_row = 4
            else:
                start_row = 3
            
            # Add DataFrame to sheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = self.default_font
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = self.border
                    
                    # Style header row
                    if r_idx == start_row:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws, df)
    
    def _add_summary_sheet(self, workbook: Workbook, message: str, filename: str):
        """Add summary/info sheet"""
        ws = workbook.create_sheet(title="Summary")
        
        ws['A1'] = f"PDF to Excel Conversion Summary"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        
        ws['A3'] = f"Original File: {filename}"
        ws['A3'].font = self.default_font
        
        ws['A4'] = f"Conversion Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'].font = self.default_font
        
        ws['A6'] = f"Status: {message}"
        ws['A6'].font = self.default_font
        
        # Adjust column width
        ws.column_dimensions['A'].width = 60
    
    def _add_error_sheet(self, workbook: Workbook, error_message: str, filename: str):
        """Add error information sheet"""
        ws = workbook.create_sheet(title="Error")
        
        ws['A1'] = "Conversion Error"
        ws['A1'].font = Font(name='Microsoft YaHei', size=12, bold=True, color='FF0000')
        ws['A1'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        ws['A3'] = f"File: {filename}"
        ws['A3'].font = self.default_font
        
        ws['A4'] = f"Error occurred at: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'].font = self.default_font
        
        ws['A6'] = "Error Details:"
        ws['A6'].font = Font(name='Microsoft YaHei', size=11, bold=True)
        
        ws['A7'] = error_message
        ws['A7'].font = self.default_font
        ws['A7'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column width
        ws.column_dimensions['A'].width = 80
    
    def _adjust_column_widths(self, worksheet, dataframe: pd.DataFrame):
        """Auto-adjust column widths based on content"""
        for col_idx, column in enumerate(dataframe.columns, 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            
            # Calculate max width needed
            max_width = len(str(column))  # Header width
            
            for value in dataframe[column]:
                if pd.notna(value):
                    # Handle Chinese characters (they take more space)
                    text = str(value)
                    # Count Chinese characters (rough estimation)
                    chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text))
                    english_chars = len(text) - chinese_chars
                    # Chinese characters are roughly 1.5 times wider
                    estimated_width = english_chars + (chinese_chars * 1.5)
                    max_width = max(max_width, estimated_width)
            
            # Set width with limits
            width = min(max_width + 2, 50)  # Max width of 50
            width = max(width, 10)  # Min width of 10
            
            worksheet.column_dimensions[column_letter].width = width
